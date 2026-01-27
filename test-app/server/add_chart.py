#!/usr/bin/env python3
import sys
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

def add_bar_chart(file_path):
    """Add a bar chart to the Summary sheet"""
    try:
        # Load the workbook
        wb = load_workbook(file_path)
        ws = wb['Summary']

        # Find the data range (excluding header and total rows)
        max_row = ws.max_row
        max_col = ws.max_column

        # Create bar chart
        chart = BarChart()
        chart.type = "col"  # Column chart (vertical bars)
        chart.style = 10
        chart.title = "Simplification Items by Workstream"
        chart.y_axis.title = "Item Count"
        chart.x_axis.title = "Workstream"

        # Data range: from column B (categories) to second-to-last column (exclude Total), rows 1 to max_row-1 (exclude Total row)
        data = Reference(ws, min_col=2, max_col=max_col-1, min_row=1, max_row=max_row-1)

        # Category labels: column A (workstreams), from row 2 to max_row-1
        cats = Reference(ws, min_col=1, min_row=2, max_row=max_row-1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Set chart size
        chart.width = 20
        chart.height = 12

        # Add chart to the sheet at position H2
        ws.add_chart(chart, "H2")

        # Save the workbook
        wb.save(file_path)
        print(f"Chart added successfully to {file_path}")
        return 0

    except Exception as e:
        print(f"Error adding chart: {str(e)}", file=sys.stderr)
        return 1

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python3 add_chart.py <excel_file_path>", file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1]
    sys.exit(add_bar_chart(file_path))
