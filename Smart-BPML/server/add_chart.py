#!/usr/bin/env python3
import sys
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

# Brand color palette (hex without #)
COLORS = [
    'FD5108',  # PwC orange
    '2C3E50',  # dark navy
    '27AE60',  # green
    '2980B9',  # blue
    'E67E22',  # amber
    '8E44AD',  # purple
    '16A085',  # teal
    'C0392B',  # red
    'F39C12',  # yellow-orange
    '7F8C8D',  # gray
]



def add_data_labels(chart, show_val=True):
    """Attach value data labels to every series in a chart."""
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = show_val
    chart.dLbls.showLegendKey = False
    chart.dLbls.showCatName = False
    chart.dLbls.showSerName = False
    chart.dLbls.showPercent = False


def set_series_colors(chart):
    """Apply brand color palette to each series."""
    for i, series in enumerate(chart.series):
        color = COLORS[i % len(COLORS)]
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.solidFill = color


def build_grouped_bar_chart(ws, max_row, max_col):
    """Grouped column chart: workstreams on X axis, count on Y axis,
    one colored series per category with a labeled legend on the right."""
    chart = BarChart()
    chart.type     = 'col'        # vertical column chart
    chart.grouping = 'clustered'
    chart.overlap  = 0
    chart.style    = 2            # clean white background

    chart.title         = 'Simplification Items by Workstream'
    chart.y_axis.title  = 'Count'
    chart.x_axis.title  = 'Workstream'

    # Y axis: integer ticks only, light horizontal gridlines
    chart.y_axis.numFmt        = '0'
    chart.y_axis.majorGridlines = None  # openpyxl default adds gridlines; set None = keep default light ones

    # Data: columns B … (last-1)  rows 1 … (last-1)
    # Row 1 = header (category names) → used as series labels in legend
    # Rows 2 … max_row-1 = data rows (exclude the Total summary row)
    # Columns 2 … max_col-1 = category columns (exclude Total column)
    data = Reference(ws, min_col=2, max_col=max_col - 1,
                     min_row=1,   max_row=max_row - 1)
    # X axis labels: column A, rows 2 … max_row-1 (workstream names)
    cats = Reference(ws, min_col=1,
                     min_row=2, max_row=max_row - 1)

    chart.add_data(data, titles_from_data=True)  # row-1 values become series names in legend
    chart.set_categories(cats)

    # Data labels: show count on top of every bar
    add_data_labels(chart)

    # Distinct colors per category series so legend is unambiguous
    set_series_colors(chart)

    # Legend on the RIGHT so workstream X labels have full width
    # Each legend entry = colored square + category name
    chart.legend.position = 'r'

    # Make chart large enough for long workstream names and a readable legend
    chart.width  = 32   # ~22 cm wide
    chart.height = 18   # ~13 cm tall

    return chart



def add_charts(file_path):
    """Main entry point: style the Summary sheet and add two enhanced charts."""
    try:
        wb = load_workbook(file_path)

        if 'Summary' not in wb.sheetnames:
            print('Summary sheet not found — skipping charts', file=sys.stderr)
            wb.save(file_path)
            return 0

        ws       = wb['Summary']
        max_row  = ws.max_row
        max_col  = ws.max_column

        if max_row < 3 or max_col < 2:
            print('Summary sheet has insufficient data — skipping charts', file=sys.stderr)
            wb.save(file_path)
            return 0

        # 1. Grouped column chart — placed to the right of the table
        chart1 = build_grouped_bar_chart(ws, max_row, max_col)
        chart1_col = get_column_letter(max_col + 2)
        ws.add_chart(chart1, f'{chart1_col}1')

        wb.save(file_path)
        print(f'Charts added successfully to {file_path}')
        return 0

    except Exception as e:
        print(f'Error adding charts: {str(e)}', file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('Usage: python3 add_chart.py <excel_file_path>', file=sys.stderr)
        sys.exit(1)

    sys.exit(add_charts(file_path=sys.argv[1]))
